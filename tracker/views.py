from django.views import View
from django.shortcuts import render, redirect

from .forms import UploadFileForm
from .models import CarStatus

import openpyxl


class MainView(View):
    def get(self, request):
        # Select Boxes
        selects_query = """
            SELECT id, name as name, strftime("%Y-%m-%d", date_time) as date
            FROM tracker_carstatus
            GROUP BY 3;
        """
        selects = CarStatus.objects.raw(selects_query)
        cars, dates = [], []
        for s in selects:
            cars.append(s.name)
            dates.append(s.date)
        cars = set(list(cars))
        # Map
        name = request.GET.get('car', None)
        date = request.GET.get('date', None)

        if name is not None and date is not None:
            map_query = """
                SELECT id, time(date_time) as time, speed as speed, distance as distance, x as x, y as y
                FROM tracker_carstatus
                WHERE name="{name}" AND strftime("%Y-%m-%d", date_time)="{date}"
                ORDER BY 2;
            """.format(name=name, date=date)
            map_ = CarStatus.objects.raw(map_query)
            map_ = [{'time': m.time, 'speed': m.speed, 'distance': m.distance, 'x': m.x, 'y': m.y} for m in map_]
            return render(request, 'main.html', {'cars': cars, 'dates': dates, 'map': map_, 'name': name})
        else:
            return render(request, 'main.html', {'cars': cars, 'dates': dates, 'map': [], 'name': []})


class FileUploadView(View):
    def get(self, request):
        form = UploadFileForm()
        return render(request, 'file_upload.html', {'form': form})

    def post(self, request):
        excel_file = request.FILES["file"]
        wb = openpyxl.load_workbook(excel_file)
        active_sheet = wb.active
        name = active_sheet.cell(1, 1).value.replace(' ', '').split(':')[1]

        target_col_idx = [5, 6, 7, 8, 9, 11, 12, 13, 14]
        int_value_idx = [0, 4, 6]
        float_value_idx = [5, 7, 8]
        for row_idx in range(3, active_sheet.max_row + 1):
            record = [active_sheet.cell(row_idx, col_idx).value for col_idx in target_col_idx]

            for i_idx in int_value_idx:
                try:
                    record[i_idx] = int(record[i_idx].split(' ')[0].replace(',', ''))
                except ValueError as E:
                    record[i_idx] = 0
            for f_idx in float_value_idx:
                try:
                    record[f_idx] = float(record[f_idx].split(' ')[0])
                except ValueError as E:
                    record[f_idx] = 0

            record = CarStatus(name=name, ping=record[0], date_time=record[1], key_on_datetime=record[2],
                               key_off_datetime=record[3], speed=record[4], distance=record[5],
                               cum_distance=record[6], x=record[7], y=record[8])
            record.save()
        return redirect('main')


class DBInitializeView(View):
    def get(self, request):
        all_records = CarStatus.objects.all()
        all_records.delete()
        return redirect('main')
